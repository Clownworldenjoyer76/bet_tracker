#!/usr/bin/env python3

import csv
import re
import os
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR = Path(__file__).parent.parent
INPUT_RAW_DIR = BASE_DIR / "docs/win/dump"
INPUT_NORM_DIR = BASE_DIR / "docs/win/manual/normalized"
FINAL_OUTPUT_DIR = BASE_DIR / "docs/win/cleaned"
FINAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Helper for team cleaning
def strip_team(name):
    if not name: return ""
    s = re.sub(r"\([^)]*\)", "", str(name))
    s = re.sub(r"\s+\d+\s*-\s*\d+\s*$", "", s)
    return s.strip()

def pct_to_decimal(value):
    if not value: return ""
    s = str(value).strip()
    return str(float(s[:-1]) / 100) if s.endswith("%") else s

# ============================================================
# EXTRACTION LOGIC
# ============================================================

def get_raw_rows(league):
    """Reads Excel files and returns list of dicts with basic cleanup."""
    files = sorted(INPUT_RAW_DIR.glob(f"{league}_*.xlsx"))
    all_extracted = []

    for path in files:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[1:] # Skip header
        
        for row in rows:
            if not row or not any(row): continue
            
            # Basic shared parsing
            dt_lines = str(row[0]).splitlines() if row[0] else []
            date = dt_lines[0] if len(dt_lines) > 0 else ""
            time = dt_lines[1] if len(dt_lines) > 1 else ""
            
            teams = str(row[1]).splitlines() if row[1] else []
            team_a = strip_team(teams[0]) if len(teams) > 0 else ""
            team_b = strip_team(teams[1]) if len(teams) > 1 else ""
            
            wins = str(row[2]).splitlines() if row[2] else []
            win_a = pct_to_decimal(wins[0]) if len(wins) > 0 else ""
            win_b = pct_to_decimal(wins[1]) if len(wins) > 1 else ""

            # League specific mapping
            data = {"date": date, "time": time, "team_a": team_a, "team_b": team_b, 
                    "win_a": win_a, "win_b": win_b, "league": league}

            if league == "soc":
                data.update({"draw": pct_to_decimal(row[3]), "goals_a": str(row[4]).splitlines()[0] if row[4] else "", 
                             "goals_b": str(row[4]).splitlines()[1] if row[4] and len(str(row[4]).splitlines())>1 else "",
                             "total": row[5] or ""})
            elif league == "nhl":
                data.update({"goals_a": str(row[3]).splitlines()[0] if row[3] else "", 
                             "goals_b": str(row[3]).splitlines()[1] if row[3] and len(str(row[3]).splitlines())>1 else "",
                             "total": row[4] or ""})
            else: # NBA / NCAAB
                data.update({"pts_a": str(row[3]).splitlines()[0] if row[3] else "", 
                             "pts_b": str(row[3]).splitlines()[1] if row[3] and len(str(row[3]).splitlines())>1 else "",
                             "total": row[4] or ""})
            
            all_extracted.append(data)
    return all_extracted

# ============================================================
# PROCESSING & MERGING
# ============================================================

def process_league_unified(league):
    raw_data = get_raw_rows(league)
    if not raw_data: return

    # 1. Determine file date (Format: YYYY_mm_dd for matching)
    search_date = ""
    for r in raw_data:
        if r['date']:
            try:
                search_date = datetime.strptime(r['date'], "%m/%d/%Y").strftime("%Y_%m_%d")
                break
            except: 
                try:
                    search_date = datetime.strptime(r['date'], "%m/%d/%y").strftime("%Y_%m_%d")
                    break
                except: continue
    
    if not search_date: return

    # 2. Load DK Manual Data safely (Force underscore match)
    dk_file = INPUT_NORM_DIR / f"norm_dk_{league}_totals_{search_date}.csv"
    dk_map = {}
    
    if dk_file.exists():
        with open(dk_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                gid = row.get("game_id")
                if not gid: continue
                if gid not in dk_map: 
                    dk_map[gid] = {"total": row.get("total", ""), "over": {}, "under": {}}
                
                side = (row.get("side") or "").lower()
                if side in ["over", "under"]:
                    dk_map[gid][side] = row
    else:
        print(f"Warning: Missing DK file {dk_file.name}")

    # 3. Group and Merge
    final_rows = []
    for i, entry in enumerate(raw_data):
        try:
            dt_obj = datetime.strptime(entry['date'], "%m/%d/%Y")
        except:
            try:
                dt_obj = datetime.strptime(entry['date'], "%m/%d/%y")
            except:
                continue
        
        date_id_str = dt_obj.strftime("%Y_%m_%d")
        gid = f"{league}_{date_id_str}_game_{i+1}"
        
        p_key_a = entry.get("pts_a") or entry.get("goals_a") or ""
        p_key_b = entry.get("pts_b") or entry.get("goals_b") or ""
        
        # Pull data from dk_map based on game_id
        dk = dk_map.get(gid, {"total": "", "over": {}, "under": {}})

        final_rows.append({
            "game_id": gid, 
            "date": entry["date"], 
            "time": entry["time"],
            "away_team": entry["team_a"], 
            "home_team": entry["team_b"],
            "away_points": p_key_a, 
            "home_points": p_key_b,
            "total_points": entry.get("total", ""), 
            "away_win_probability": entry["win_a"],
            "home__win_probability": entry["win_b"], 
            "total": dk["total"],
            "over_total_odds": dk["over"].get("odds", ""), 
            "under_total_odds": dk["under"].get("odds", ""),
            "over_handle_pct": dk["over"].get("handle_pct", ""), 
            "under_handle_pct": dk["under"].get("handle_pct", ""),
            "over_bets_pct": dk["over"].get("bets_pct", ""), 
            "under_bets_pct": dk["under"].get("bets_pct", ""),
            "league": league
        })

    # 4. Write Final Output
    if final_rows:
        # Save output using dashes for clean filename: clean_league_YYYY-mm-dd.csv
        out_date = search_date.replace("_", "-")
        out_path = FINAL_OUTPUT_DIR / f"clean_{league}_{out_date}.csv"
        headers = [
            "game_id", "date", "time", "away_team", "home_team",
            "away_points", "home_points", "total_points",
            "away_win_probability", "home__win_probability",
            "total", "over_total_odds", "under_total_odds",
            "over_handle_pct", "under_handle_pct", "over_bets_pct",
            "under_bets_pct", "league"
        ]
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(final_rows)
        print(f"Successfully created: {out_path.name}")

def main():
    for league in ["nba", "ncaab", "nhl", "soc"]:
        process_league_unified(league)

if __name__ == "__main__":
    main()
