#!/usr/bin/env python3

import csv
import re
from pathlib import Path
from openpyxl import load_workbook

# Config
BASE_DIR = Path(__file__).parent.parent
INPUT_RAW_DIR = BASE_DIR / "docs/win/dump"
INPUT_NORM_DIR = BASE_DIR / "docs/win/manual/normalized"
FINAL_OUTPUT_DIR = BASE_DIR / "docs/win/cleaned"
FINAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def strip_team(name):
    if not name: return ""
    s = re.sub(r"\([^)]*\)", "", str(name))
    s = re.sub(r"\s+\d+\s*-\s*\d+\s*$", "", s)
    return s.strip()

def pct_to_decimal(value):
    if not value: return ""
    s = str(value).strip()
    return str(float(s[:-1]) / 100) if s.endswith("%") else s

def run_clean():
    for league in ["nba", "ncaab", "nhl", "soc"]:
        for raw_file in INPUT_RAW_DIR.glob(f"{league}_*.xlsx"):
            # Date extraction for file matching
            date_match = re.search(r"(\d{4}-\d{2}-\d{2})", raw_file.name)
            if not date_match: continue
            
            date_dash = date_match.group(1)
            date_underscore = date_dash.replace("-", "_")

            # 1. Build the Lookup Table from Normalized CSV
            dk_data = {} # Key: game_id
            dk_csv = INPUT_NORM_DIR / f"norm_dk_{league}_totals_{date_underscore}.csv"
            
            if dk_csv.exists():
                with open(dk_csv, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        gid = row.get("game_id")
                        side = str(row.get("side", "")).upper().strip()
                        
                        if gid not in dk_data:
                            dk_data[gid] = {"total": row.get("total", ""), "OVER": {}, "UNDER": {}}
                        
                        if side in ["OVER", "UNDER"]:
                            dk_data[gid][side] = row

            # 2. Process Raw Excel
            wb = load_workbook(raw_file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))[1:]
            
            cleaned_rows = []
            for i, row in enumerate(rows):
                if not row or not any(row): continue
                
                gid = f"{league}_{date_underscore}_game_{i+1}"
                dk = dk_data.get(gid, {"total": "", "OVER": {}, "UNDER": {}})
                
                # Parse layout
                dt = str(row[0]).splitlines()
                tm = str(row[1]).splitlines()
                wn = str(row[2]).splitlines()
                idx = 4 if league == "soc" else 3
                pt = str(row[idx]).splitlines() if row[idx] else ["", ""]

                cleaned_rows.append({
                    "game_id": gid,
                    "date": dt[0] if dt else "",
                    "time": dt[1] if len(dt) > 1 else "",
                    "away_team": strip_team(tm[0]) if tm else "",
                    "home_team": strip_team(tm[1]) if len(tm) > 1 else "",
                    "away_points": pt[0] if pt else "",
                    "home_points": pt[1] if len(pt) > 1 else "",
                    "total_points": row[idx + 1] if len(row) > idx + 1 else "",
                    "away_win_probability": pct_to_decimal(wn[0]) if wn else "",
                    "home__win_probability": pct_to_decimal(wn[1]) if len(wn) > 1 else "",
                    # The requested mapping
                    "total": dk.get("total", ""),
                    "over_total_odds": dk["OVER"].get("odds", ""),
                    "under_total_odds": dk["UNDER"].get("odds", ""),
                    "over_handle_pct": dk["OVER"].get("handle_pct", ""),
                    "under_handle_pct": dk["UNDER"].get("handle_pct", ""),
                    "over_bets_pct": dk["OVER"].get("bets_pct", ""),
                    "under_bets_pct": dk["UNDER"].get("bets_pct", ""),
                    "league": league
                })

            # 3. Save
            if cleaned_rows:
                output_file = FINAL_OUTPUT_DIR / f"clean_{league}_{date_dash}.csv"
                fields = [
                    "game_id", "date", "time", "away_team", "home_team",
                    "away_points", "home_points", "total_points",
                    "away_win_probability", "home__win_probability",
                    "total", "over_total_odds", "under_total_odds",
                    "over_handle_pct", "under_handle_pct", "over_bets_pct",
                    "under_bets_pct", "league"
                ]
                with open(output_file, "w", newline="", encoding="utf-8") as f:
                    writer = csv.DictWriter(f, fieldnames=fields)
                    writer.writeheader()
                    writer.writerows(cleaned_rows)

if __name__ == "__main__":
    run_clean()
