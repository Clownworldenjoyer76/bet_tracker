#!/usr/bin/env python3

import csv
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR = Path(__file__).parent.parent
INPUT_RAW_DIR = BASE_DIR / "docs/win/dump"
INPUT_NORM_DIR = BASE_DIR / "docs/win/manual/normalized"
TEMP_CLEANED_DIR = BASE_DIR / "docs/win/cleaned"
FINAL_OUTPUT_DIR = BASE_DIR / "docs/win/cleaned/final"

# Ensure directories exist
TEMP_CLEANED_DIR.mkdir(parents=True, exist_ok=True)
FINAL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# PART 1: EXTRACT FROM EXCEL DUMPS
# ============================================================

def strip_team(name):
    if not name: return ""
    s = re.sub(r"\([^)]*\)", "", str(name))
    s = re.sub(r"\s+\d+\s*-\s*\d+\s*$", "", s)
    return s.strip()

def pct_to_decimal(value):
    if not value: return ""
    s = str(value).strip()
    return str(float(s[:-1]) / 100) if s.endswith("%") else s

def get_raw_rows(league):
    files = sorted(INPUT_RAW_DIR.glob(f"{league}_*.xlsx"))
    all_extracted = []
    for path in files:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))[1:] 
        for row in rows:
            if not row or not any(row): continue
            dt_lines = str(row[0]).splitlines() if row[0] else []
            teams = str(row[1]).splitlines() if row[1] else []
            wins = str(row[2]).splitlines() if row[2] else []
            data = {
                "date": dt_lines[0] if dt_lines else "",
                "time": dt_lines[1] if len(dt_lines) > 1 else "",
                "team_a": strip_team(teams[0]) if teams else "",
                "team_b": strip_team(teams[1]) if len(teams) > 1 else "",
                "win_a": pct_to_decimal(wins[0]) if wins else "",
                "win_b": pct_to_decimal(wins[1]) if len(wins) > 1 else "",
                "league": league
            }
            # Handle sport-specific column offsets
            if league == "soc":
                data.update({"goals_a": str(row[4]).splitlines()[0] if row[4] else "", 
                             "goals_b": str(row[4]).splitlines()[1] if row[4] and len(str(row[4]).splitlines())>1 else "",
                             "total_pts": row[5] or ""})
            elif league == "nhl":
                data.update({"goals_a": str(row[3]).splitlines()[0] if row[3] else "", 
                             "goals_b": str(row[3]).splitlines()[1] if row[3] and len(str(row[3]).splitlines())>1 else "",
                             "total_pts": row[4] or ""})
            else:
                data.update({"pts_a": str(row[3]).splitlines()[0] if row[3] else "", 
                             "pts_b": str(row[3]).splitlines()[1] if row[3] and len(str(row[3]).splitlines())>1 else "",
                             "total_pts": row[4] or ""})
            all_extracted.append(data)
    return all_extracted

def process_initial_cleaning():
    """Converts Excel dumps to initial CSVs"""
    for lg in ["nba", "ncaab", "nhl", "soc"]:
        raw_data = get_raw_rows(lg)
        if not raw_data: continue

        # Determine file date from the first row
        search_date = ""
        for r in raw_data:
            if r['date']:
                for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                    try:
                        search_date = datetime.strptime(r['date'], fmt).strftime("%Y_%m_%d")
                        break
                    except: continue
                if search_date: break
        
        if not search_date: continue

        final_rows = []
        for i, entry in enumerate(raw_data):
            gid = f"{lg}_{search_date}_game_{i+1}"
            final_rows.append({
                "game_id": gid, "date": entry["date"], "time": entry["time"],
                "away_team": entry["team_a"], "home_team": entry["team_b"],
                "away_points": entry.get("pts_a") or entry.get("goals_a") or "",
                "home_points": entry.get("pts_b") or entry.get("goals_b") or "",
                "total_points": entry.get("total_pts", ""),
                "away_win_probability": entry["win_a"],
                "home__win_probability": entry["win_b"],
                "total": "", "over_total_odds": "", "under_total_odds": "",
                "over_handle_pct": "", "under_handle_pct": "",
                "over_bets_pct": "", "under_bets_pct": "", "league": lg
            })

        out_date = search_date.replace("_", "-")
        out_path = TEMP_CLEANED_DIR / f"clean_{lg}_{out_date}.csv"
        headers = ["game_id", "date", "time", "away_team", "home_team", "away_points", "home_points", "total_points", "away_win_probability", "home__win_probability", "total", "over_total_odds", "under_total_odds", "over_handle_pct", "under_handle_pct", "over_bets_pct", "under_bets_pct", "league"]
        
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(final_rows)
        print(f"Step 1 Complete: Created {out_path.name}")

# ============================================================
# PART 2: TEAM-BASED LOOKUP FOR TOTALS DATA
# ============================================================

def run_fix_and_finalize():
    """Matches data by team name and moves to final folder"""
    for file_path in TEMP_CLEANED_DIR.glob("clean_*.csv"):
        if "final" in file_path.parts: continue
        
        match = re.search(r"clean_([a-z]+)_(\d{4})[-_](\d{2})[-_](\d{2})", file_path.name)
        if not match: continue
        
        league, yyyy, mm, dd = match.groups()
        date_str = f"{yyyy}_{mm}_{dd}"
        
        norm_filename = f"norm_dk_{league}_totals_{date_str}.csv"
        norm_path = INPUT_NORM_DIR / norm_filename
        
        if not norm_path.exists():
            print(f"DraftKings File Not Found: {norm_filename}")
            continue

        # Map Normalized Totals Data by TEAM NAME
        team_lookup = {}
        with open(norm_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                team = str(row.get("team", "")).strip().lower()
                side = str(row.get("side", "")).strip().upper()
                if team not in team_lookup:
                    team_lookup[team] = {"total": row.get("total", ""), "OVER": {}, "UNDER": {}}
                if side in ["OVER", "UNDER"]:
                    team_lookup[team][side] = row

        # Update and Save to Final
        updated_rows = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            for row in reader:
                away_team = str(row.get("away_team", "")).strip().lower()
                dk = team_lookup.get(away_team)
                if dk:
                    row["total"] = dk.get("total", "")
                    row["over_total_odds"] = dk["OVER"].get("odds", "")
                    row["under_total_odds"] = dk["UNDER"].get("odds", "")
                    row["over_handle_pct"] = dk["OVER"].get("handle_pct", "")
                    row["under_handle_pct"] = dk["UNDER"].get("handle_pct", "")
                    row["over_bets_pct"] = dk["OVER"].get("bets_pct", "")
                    row["under_bets_pct"] = dk["UNDER"].get("bets_pct", "")
                updated_rows.append(row)

        out_path = FINAL_OUTPUT_DIR / file_path.name
        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(updated_rows)
        print(f"Step 2 Complete: Finalized {out_path.name}")

if __name__ == "__main__":
    process_initial_cleaning()
    run_fix_and_finalize()
