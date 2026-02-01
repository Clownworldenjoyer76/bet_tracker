#!/usr/bin/env python3

import csv
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# ============================================================
# Shared paths
# ============================================================

INPUT_DIR = Path("docs/win/dump")
OUTPUT_DIR = Path("docs/win/clean/step_1/")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ============================================================
# ======================= SOCCER =============================
# ============================================================

SOCCER_LEAGUE = "soc"

SOCCER_HEADERS = [
    "date",
    "time",
    "team",
    "opponent",
    "goals",
    "total_goals",
    "win_probability",
    "draw_probability",
    "best_ou",
    "bet_type",
    "league",
]


def strip_team(name):
    if name is None:
        return ""
    s = str(name)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"\s+\d+\s*-\s*\d+\s*$", "", s)
    return s.strip()


def pct_to_decimal(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith("%"):
        return str(float(s[:-1]) / 100)
    return s


def parse_best_ou_soccer(value):
    if value is None:
        return ""
    m = re.search(r"(\d+)", str(value))
    if not m:
        return ""
    return f"{m.group(1)}.5"


def run_soccer():
    files = sorted(INPUT_DIR.glob("soc_*.xlsx"))
    if not files:
        return

    for path in files:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        data_rows = rows[1:]
        output_rows = []

        for row in data_rows:
            if not row or (not row[1] and not row[2] and not row[4]):
                continue

            dt_lines = str(row[0]).splitlines() if row[0] else []
            date = dt_lines[0] if len(dt_lines) > 0 else ""
            time = dt_lines[1] if len(dt_lines) > 1 else ""

            teams = str(row[1]).splitlines() if row[1] else []
            team_a = strip_team(teams[0]) if len(teams) > 0 else ""
            team_b = strip_team(teams[1]) if len(teams) > 1 else ""

            wins = str(row[2]).splitlines() if row[2] else []
            win_a = pct_to_decimal(wins[0]) if len(wins) > 0 else ""
            win_b = pct_to_decimal(wins[1]) if len(wins) > 1 else ""

            draw = pct_to_decimal(row[3]) if len(row) > 3 else ""

            goals = str(row[4]).splitlines() if len(row) > 4 and row[4] else []
            goals_a = goals[0] if len(goals) > 0 else ""
            goals_b = goals[1] if len(goals) > 1 else ""

            total_goals = row[5] if len(row) > 5 and row[5] is not None else ""

            best_ou_raw = row[6] if len(row) > 6 else ""
            best_ou = parse_best_ou_soccer(best_ou_raw)

            output_rows.append([
                date, time,
                team_a, team_b,
                goals_a, total_goals,
                win_a, draw,
                best_ou,
                "win",
                SOCCER_LEAGUE
            ])

            output_rows.append([
                date, time,
                team_b, team_a,
                goals_b, total_goals,
                win_b, draw,
                best_ou,
                "win",
                SOCCER_LEAGUE
            ])

            output_rows.append([
                date, time,
                "DRAW", f"{team_a} vs {team_b}",
                "", total_goals,
                "", draw,
                best_ou,
                "draw",
                SOCCER_LEAGUE
            ])

        if not output_rows:
            continue

        file_date = datetime.strptime(
            data_rows[0][0].splitlines()[0],
            "%m/%d/%Y"
        ).strftime("%Y-%m-%d")

        out_path = OUTPUT_DIR / f"win_prob__clean_{SOCCER_LEAGUE}_{file_date}.csv"

        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(SOCCER_HEADERS)
            writer.writerows(output_rows)

# ============================================================
# ======================= NHL ================================
# ============================================================

NHL_LEAGUE = "nhl"

NHL_HEADERS = [
    "date",
    "time",
    "team",
    "opponent",
    "goals",
    "total_goals",
    "win_probability",
    "best_ou",
    "league",
]


def parse_best_ou_nhl(value):
    if value is None:
        return ""
    nums = re.findall(r"\d+\.?\d*", str(value).lower())
    if not nums:
        return ""
    return f"{int(float(nums[0]))}.5"


def run_nhl():
    files = sorted(INPUT_DIR.glob("nhl_*.xlsx"))
    if not files:
        return

    for path in files:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        data_rows = rows[1:]
        output_rows = []
        file_date = ""

        for row in data_rows:
            if not any(row):
                continue

            dt_lines = str(row[0]).splitlines() if row[0] else []
            date = dt_lines[0] if len(dt_lines) > 0 else ""
            time = dt_lines[1] if len(dt_lines) > 1 else ""

            if date and not file_date:
                file_date = datetime.strptime(date, "%m/%d/%Y").strftime("%Y-%m-%d")

            teams = str(row[1]).splitlines() if row[1] else []
            team_a = strip_team(teams[0]) if len(teams) > 0 else ""
            team_b = strip_team(teams[1]) if len(teams) > 1 else ""

            wins = str(row[2]).splitlines() if row[2] else []
            win_a = pct_to_decimal(wins[0]) if len(wins) > 0 else ""
            win_b = pct_to_decimal(wins[1]) if len(wins) > 1 else ""

            goals = str(row[3]).splitlines() if row[3] else []
            goals_a = goals[0] if len(goals) > 0 else ""
            goals_b = goals[1] if len(goals) > 1 else ""

            total_goals = row[4] if row[4] is not None else ""
            best_ou = parse_best_ou_nhl(row[5]) if len(row) > 5 else ""

            output_rows.append([
                date, time,
                team_a, team_b,
                goals_a, total_goals,
                win_a, best_ou,
                NHL_LEAGUE
            ])

            output_rows.append([
                date, time,
                team_b, team_a,
                goals_b, total_goals,
                win_b, best_ou,
                NHL_LEAGUE
            ])

        if output_rows:
            out_path = OUTPUT_DIR / f"win_prob__clean_{NHL_LEAGUE}_{file_date}.csv"
            with out_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(NHL_HEADERS)
                writer.writerows(output_rows)

# ============================================================
# ======================= NBA ================================
# ============================================================

NBA_LEAGUE = "nba"

NBA_HEADERS = [
    "date",
    "time",
    "team",
    "opponent",
    "points",
    "total_points",
    "win_probability",
    "best_ou",
    "league",
]


def parse_best_ou_nba(raw):
    if not raw:
        return ""
    s = str(raw).strip().lower()
    if not (s.startswith("o") or s.startswith("u")):
        return ""
    s = s[1:].split("-")[0].replace("½", ".5")
    return f"{int(float(s))}.5"


def run_nba():
    files = sorted(INPUT_DIR.glob("nba_*.xlsx"))
    if not files:
        return

    for path in files:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        data_rows = rows[1:]
        output_rows = []
        file_date = ""

        for row in data_rows:
            if not row or not row[0] or not row[1]:
                continue

            dt_lines = str(row[0]).splitlines()
            if len(dt_lines) < 2:
                continue

            date, time = dt_lines[0], dt_lines[1]

            if not file_date:
                file_date = datetime.strptime(date, "%m/%d/%Y").strftime("%Y-%m-%d")

            teams = str(row[1]).splitlines()
            team_a = strip_team(teams[0])
            team_b = strip_team(teams[1])

            wins = str(row[2]).splitlines()
            win_a = pct_to_decimal(wins[0])
            win_b = pct_to_decimal(wins[1])

            points = str(row[3]).splitlines() if row[3] else []
            pts_a = points[0] if len(points) > 0 else ""
            pts_b = points[1] if len(points) > 1 else ""

            total_points = row[4] or ""
            best_ou = parse_best_ou_nba(row[5]) if len(row) > 5 else ""

            output_rows.append([
                date, time,
                team_a, team_b,
                pts_a, total_points,
                win_a, best_ou,
                NBA_LEAGUE
            ])

            output_rows.append([
                date, time,
                team_b, team_a,
                pts_b, total_points,
                win_b, best_ou,
                NBA_LEAGUE
            ])

        if output_rows:
            out_path = OUTPUT_DIR / f"win_prob__clean_{NBA_LEAGUE}_{file_date}.csv"
            with out_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(NBA_HEADERS)
                writer.writerows(output_rows)

# ============================================================
# ======================= NCAAB ==============================
# ============================================================

NCAAB_LEAGUE = "ncaab"
NCAAB_HEADERS = NBA_HEADERS


def parse_best_ou_ncaab(value):
    nums = re.findall(r"\d+", str(value)) if value else []
    return f"{nums[0]}.5" if nums else ""


def round_prob(value):
    try:
        return f"{round(float(value), 2):.2f}"
    except Exception:
        return ""


def run_ncaab():
    files = sorted(INPUT_DIR.glob("ncaab_*.xlsx"))
    if not files:
        return

    for path in files:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        data_rows = rows[1:]
        output_rows = []
        file_date = ""

        for row in data_rows:
            if not row or not row[0] or not row[1]:
                continue

            dt_lines = str(row[0]).splitlines()
            if len(dt_lines) < 2:
                continue

            date, time = dt_lines[0], dt_lines[1]

            if not file_date:
                file_date = datetime.strptime(date, "%m/%d/%Y").strftime("%Y_%m_%d")

            teams = str(row[1]).splitlines()
            team_a = strip_team(teams[0])
            team_b = strip_team(teams[1])

            wins = str(row[2]).splitlines()
            win_a = round_prob(pct_to_decimal(wins[0]))
            win_b = round_prob(pct_to_decimal(wins[1]))

            points = str(row[3]).splitlines() if row[3] else []
            pts_a = points[0] if len(points) > 0 else ""
            pts_b = points[1] if len(points) > 1 else ""

            total_points = row[4] or ""
            best_ou = parse_best_ou_ncaab(row[5]) if len(row) > 5 else ""

            output_rows.append([
                date, time,
                team_a, team_b,
                pts_a, total_points,
                win_a, best_ou,
                NCAAB_LEAGUE
            ])

            output_rows.append([
                date, time,
                team_b, team_a,
                pts_b, total_points,
                win_b, best_ou,
                NCAAB_LEAGUE
            ])

        if output_rows:
            out_path = OUTPUT_DIR / f"win_prob__clean_{NCAAB_LEAGUE}_{file_date}.csv"
            with out_path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(NCAAB_HEADERS)
                writer.writerows(output_rows)

# ============================================================
# ======================= GAME ID PASS =======================
# ============================================================

PATTERN = "docs/win/clean/win_prob__clean_*.csv"


def is_empty(value):
    return value is None or value == ""


def process_file(path: Path):
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = list(reader.fieldnames or [])

    if "game_id" not in fieldnames:
        fieldnames.append("game_id")
        for row in rows:
            row["game_id"] = ""

    game_counter = 1
    total_rows = len(rows)

    # Pass 1: original pairing logic
    for i in range(total_rows):
        row_a = rows[i]
        if not is_empty(row_a.get("game_id")):
            continue

        team_a = row_a.get("team")
        opp_a = row_a.get("opponent")

        for j in range(i + 1, total_rows):
            row_b = rows[j]
            if not is_empty(row_b.get("game_id")):
                continue

            if row_b.get("team") == opp_a and row_b.get("opponent") == team_a:
                league = row_a.get("league")
                raw_date = row_a.get("date")

                try:
                    dt = datetime.strptime(raw_date, "%m/%d/%Y")
                except ValueError:
                    dt = datetime.strptime(raw_date, "%m/%d/%y")

                formatted_date = dt.strftime("%Y_%m_%d")
                game_id = f"{league}_{formatted_date}_game_{game_counter}"

                row_a["game_id"] = game_id
                row_b["game_id"] = game_id
                game_counter += 1
                break

    # Pass 2: attach DRAW rows
    for row in rows:
        if row.get("team") != "DRAW":
            continue
        if not is_empty(row.get("game_id")):
            continue

        opp = row.get("opponent")
        if not opp or " vs " not in opp:
            continue

        team_a, team_b = [x.strip() for x in opp.split(" vs ", 1)]

        for other in rows:
            if is_empty(other.get("game_id")):
                continue

            if (
                (other.get("team") == team_a and other.get("opponent") == team_b) or
                (other.get("team") == team_b and other.get("opponent") == team_a)
            ):
                row["game_id"] = other["game_id"]
                break

    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def add_game_ids_to_all_clean_csvs():
    for path in sorted(Path().glob(PATTERN)):
        process_file(path)

# ============================================================
# ======================= MAIN ================================
# ============================================================

def main():
    run_soccer()
    run_nhl()
    run_nba()
    run_ncaab()
    add_game_ids_to_all_clean_csvs()


if __name__ == "__main__":
    main()
