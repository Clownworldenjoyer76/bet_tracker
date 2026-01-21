#!/usr/bin/env python3

import csv
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

INPUT_DIR = Path("docs/win/dump")
OUTPUT_DIR = Path("docs/win/clean")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

LEAGUE = "soc"

HEADERS = [
    "date",
    "time",
    "team",
    "opponent",
    "goals",
    "total_goals",
    "win_probability",
    "draw_probability",
    "best_ou",
    "league",
]


def strip_team(name):
    if name is None:
        return ""
    s = str(name)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"\s+\d+\s*-\s*\d+\s*$", "", s)
    return s.strip()


def pct_to_decimal(value):
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith("%"):
        return str(float(s[:-1]) / 100)
    return s


def parse_best_ou(value):
    if value is None:
        return ""
    m = re.search(r"(\d+)", str(value))
    if not m:
        return ""
    return f"{m.group(1)}.5"


def main():
    files = sorted(INPUT_DIR.glob("soc_*.xlsx"))
    if not files:
        raise RuntimeError("No input files found")

    for path in files:
        wb = load_workbook(path, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))

        # row 0 = header only
        data_rows = rows[1:]

        output_rows = []

        for row in data_rows:
            # SKIP rows with no visible game data
            if not row[1] and not row[2] and not row[5]:
                continue

            dt_lines = str(row[0]).splitlines() if row[0] else []
            date = dt_lines[0] if len(dt_lines) > 0 else ""
            time = dt_lines[1] if len(dt_lines) > 1 else ""

            teams = str(row[1]).splitlines() if row[1] else []
            team_a = strip_team(teams[0]) if len(teams) > 0 else ""
            team_b = strip_team(teams[1]) if len(teams) > 1 else ""

            wins = str(row[2]).splitlines() if row[2] else []
            win_a = pct_to_decimal(wins[0]) if len(wins) > 0 else ""
            win_b = pct_to_decimal(wins[1]) if len(wins) > 1 else ""

            draw = pct_to_decimal(row[3])

            goals = str(row[5]).splitlines() if row[5] else []
            goals_a = goals[0] if len(goals) > 0 else ""
            goals_b = goals[1] if len(goals) > 1 else ""

            total_goals = row[6] if row[6] is not None else ""
            best_ou = parse_best_ou(row[7])

            output_rows.append([
                date, time,
                team_a, team_b,
                goals_a, total_goals,
                win_a, draw,
                best_ou, LEAGUE
            ])

            output_rows.append([
                date, time,
                team_b, team_a,
                goals_b, total_goals,
                win_b, draw,
                best_ou, LEAGUE
            ])

        file_date = datetime.strptime(
            data_rows[0][0].splitlines()[0],
            "%m/%d/%Y"
        ).strftime("%Y-%m-%d")

        out_path = OUTPUT_DIR / f"win_prob__clean_{LEAGUE}_{file_date}.csv"

        with out_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(HEADERS)
            writer.writerows(output_rows)


if __name__ == "__main__":
    main()
